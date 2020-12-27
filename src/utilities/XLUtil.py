import openpyxl

from src.utilities import genericUtil


class XLUtil:
    def __init__(self, file_path, sheet_name):
        self.file = file_path
        self.book = openpyxl.load_workbook(self.file)
        self.sheet = self.book[sheet_name]

    def get_rows_count(self):
        return self.sheet.max_row

    def get_columns_count(self):
        return self.sheet.max_column

    def read_data(self, row_index, column_index):
        return self.sheet.cell(row=row_index, column=column_index).value

    def write_data(self, row_index, column_index, data):
        self.sheet.cell(row=row_index, column=column_index).value = data
        self.book.save(self.file)

    def read_data_of_random_row(self):
        row_data = {}
        row = genericUtil.generate_random_number(self.get_rows_count(), 2)
        for i in range(1, self.get_columns_count() + 1):
            row_data[self.sheet.cell(1, i).value] = self.sheet.cell(row, i).value
        #   row_data={'header_column':'cell_value'}
        return row_data
